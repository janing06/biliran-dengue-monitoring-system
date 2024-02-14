from django.shortcuts import render, redirect
from .models import *
from django.http import HttpResponse, JsonResponse
import json

from django.conf import settings
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek, TruncQuarter, TruncYear
import geojson

import csv
import os
import numpy as np

from datetime import datetime, timedelta

import pandas as pd
from pmdarima.arima import auto_arima
from statsmodels.tsa.statespace.sarimax import SARIMAX
from pmdarima import auto_arima
from sklearn.model_selection import train_test_split
import warnings

from django.db.models import Q

from django.core.paginator import Paginator

from django.contrib.auth.decorators import login_required, user_passes_test

from django.db.models.functions import Coalesce
from django.db.models import Count, OuterRef, Subquery, Sum
from django.db.models.functions import Coalesce, TruncDate
from django.core import serializers

from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa

from django.contrib import messages

import openpyxl
from openpyxl.styles import Font

from django.shortcuts import get_object_or_404

from django.core.exceptions import ObjectDoesNotExist

def superuser_check(user):
    return user.is_superuser


@login_required(login_url='/admin/login/')
@user_passes_test(superuser_check)
def download_csv(request):
    
    currentDate = datetime.now().date() + timedelta(days=1)
    startofYear = datetime(currentDate.year, 1, 1).date()
    
    municipals = Municipal.objects.all()

    return render(request, 'dengue/download_csv.html',{
        'municipals': municipals,
        'fromDate': str(startofYear),
        'toDate': str(currentDate - timedelta(days=1)),    
    })


@login_required(login_url='/admin/login/')
@user_passes_test(superuser_check)
def render_to_pdf(request):
    
    selected_municipal = request.GET.get('municipal')
    selected_barangay = request.GET.get('barangay')
    fromDate = request.GET.get('fromDate')
    toDate = request.GET.get('toDate')
    
    start_date = datetime.strptime(fromDate, '%Y-%m-%d')
    end_date = datetime.strptime(toDate, '%Y-%m-%d') + timedelta(days=1)



    
    if selected_municipal == 'All':
            cases = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(date__gte=start_date, date__lte=end_date).order_by('-date')

            data_day = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncDate('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
            data_week = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncWeek('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
            data_month = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncMonth('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')

            context = {
                'cases': cases,
                'fromDate': fromDate,
                'toDate': toDate,
                'selected_municipal': selected_municipal,
                'selected_barangay': selected_barangay, 
                'data_day': data_day,
                'data_week': data_week,
                'data_month': data_month,
                'name': f"{request.user.first_name.title()} {request.user.last_name.title()}",
                
            }
    else:
        if selected_barangay == 'All':
            cases = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__municipal__code=selected_municipal).filter(date__gte=start_date, date__lte=end_date).order_by('-date')
            municipal = Municipal.objects.get(code=selected_municipal)

            data_day = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__municipal__code=selected_municipal).filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncDate('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
            data_week = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__municipal__code=selected_municipal).filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncWeek('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
            data_month = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__municipal__code=selected_municipal).filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncMonth('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')


            context = {
                'cases': cases,
                'fromDate': fromDate,
                'toDate': toDate,
                'selected_municipal': selected_municipal,
                'selected_barangay': selected_barangay, 
                'municipal': municipal,
                'data_day': data_day,
                'data_week': data_week,
                'data_month': data_month,
                'name': f"{request.user.first_name.title()} {request.user.last_name.title()}",
                
            }
        else:
            cases = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__barangay__code=selected_barangay).filter(date__gte=start_date, date__lte=end_date).order_by('-date')
            municipal = Municipal.objects.get(code=selected_municipal)
            barangay = Barangay.objects.get(code=selected_barangay)

            data_day = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__barangay__code=selected_barangay).filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncDate('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
            data_week = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__barangay__code=selected_barangay).filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncWeek('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
            data_month = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__barangay__code=selected_barangay).filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncMonth('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')


            context = {
                'cases': cases,
                'fromDate': fromDate,
                'toDate': toDate,
                'selected_municipal': selected_municipal,
                'selected_barangay': selected_barangay, 
                'municipal': municipal,
                'barangay': barangay,
                'data_day': data_day,
                'data_week': data_week,
                'data_month': data_month,
                'name': f"{request.user.first_name.title()} {request.user.last_name.title()}",
            }
            
    
    template_path = 'dengue/render_pdf_view.html'

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="report.pdf"'
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(
        html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def barangay_api(request, municipal_code):
    
    barangays = Barangay.objects.filter(municipal__code=municipal_code).values('id', 'barangay','code').order_by('barangay')
    
    return JsonResponse(list(barangays), safe=False)


@login_required(login_url='/admin/login/')
@user_passes_test(superuser_check)
def cases_csv(request):
    try:
        
        selected_municipal = request.GET.get('municipal')
        selected_barangay = request.GET.get('barangay')
        from_date = request.GET.get('fromDate')
        to_date = request.GET.get('toDate')
        
        start_date = datetime.strptime(from_date, '%Y-%m-%d')
        end_date = datetime.strptime(to_date, '%Y-%m-%d')
                
        if selected_municipal != 'All':
            if selected_barangay == 'All':
                print('print municipal')
                municipal_q = Municipal.objects.get(code=selected_municipal)
                municipal = municipal_q.municipal
                barangay = 'All'
                cases_by_date = Case.objects.filter(resident__municipal__code=selected_municipal).filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncDate('date')).annotate(total_cases=Count('id')).values('total_cases','date_only').order_by('date_only')
                
            else:
                barangay_q = Barangay.objects.get(code=selected_barangay)
                barangay = barangay_q.barangay
                municipal_q = Municipal.objects.get(code=selected_municipal)
                municipal = municipal_q.municipal
                cases_by_date = Case.objects.filter(resident__barangay__code=selected_barangay).filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncDate('date')).annotate(total_cases=Count('id')).values('total_cases','date_only').order_by('date_only')
                print('print barangay')
        else:
            print('print all')
            barangay = 'All'
            municipal = 'All'
            cases_by_date = Case.objects.filter(date__gte=start_date, date__lte=end_date).annotate(date_only=TruncDate('date')).annotate(total_cases=Count('id')).values('total_cases','date_only').order_by('date_only')

        
        df_dengue = pd.DataFrame.from_records(cases_by_date)
        df_dengue['date_only'] = pd.to_datetime(df_dengue['date_only'])
        df_dengue.set_index('date_only', inplace=True)
        daily_data = df_dengue.resample('D').sum()
        daily_data = daily_data.reindex(pd.date_range(start=start_date.strftime('%Y-%m-%d'), end=end_date.strftime('%Y-%m-%d')), fill_value=0)

        print('total cases', daily_data.sum())
        
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        title_cell = sheet.cell(row=1, column=1, value='Dengue Cases Report')
        title_cell.font = Font(size=12, bold=True)
        title_cell = sheet.cell(row=2, column=1, value=f'Municipality: {municipal}')
        title_cell.font = Font(size=12, bold=False)  
        title_cell = sheet.cell(row=3, column=1, value=f'Barangay: {barangay}')
        title_cell.font = Font(size=12, bold=False)
        title_cell = sheet.cell(row=4, column=1, value='')
        title_cell = sheet.cell(row=5, column=1, value=f'Start Date: {from_date}')
        title_cell.font = Font(size=12, bold=False)
        title_cell = sheet.cell(row=6, column=1, value=f'End Date: {to_date}')
        title_cell.font = Font(size=12, bold=False)
        title_cell = sheet.cell(row=7, column=1, value='')
        title_cell = sheet.cell(row=8, column=1, value=f'Prepared by: {request.user.first_name.title()} {request.user.last_name.title()}')
        title_cell.font = Font(size=12, bold=False)
        title_cell = sheet.cell(row=9, column=1, value=f'Prepared date: {datetime.now().strftime("%B %d, %Y")}')
        title_cell.font = Font(size=12, bold=False)
        title_cell = sheet.cell(row=10, column=1, value='')
        title_cell.font = Font(size=12, bold=False)
        title_cell = sheet.cell(row=11, column=1, value=f'Total: {daily_data["total_cases"].sum()}')
        title_cell.font = Font(size=12, bold=False)
        title_cell = sheet.cell(row=12, column=1, value='')

        
        sheet.merge_cells('A1:B1')

        header = ['Date', 'Cases']
        sheet.append(header)

        for index, row in daily_data.iterrows():
            sheet.append([index.strftime('%Y-%m-%d'), row['total_cases']])

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 10

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="DengueCases.xlsx"'

        workbook.save(response)

        return response
    
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        messages.success(request, 'No records found')
        return redirect('download_csv')


@login_required(login_url='/admin/login/')
@user_passes_test(superuser_check)
def case_geojson(request):
    
    cases = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').all().order_by('date')
      
    features = []

    this_year = datetime.now().date().year

    for case in cases:
        feature = {
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [ case.resident.geom.x,case.resident.geom.y],
            },
            "properties": {
                "date": case.date.strftime('%Y-%m-%d'),
                "dengue_type": case.dengue_type,
                "resident_id": case.resident.resident_id,
                "municipal": case.resident.municipal.municipal,
                "barangay": case.resident.barangay.barangay,
                "sex": case.resident.sex,
                "age": this_year - int(case.resident.birth_date.strftime('%Y'))
            },
        }
        features.append(feature)

    geojson = {
        "type": "FeatureCollection",
        "features": features,
    }
    

    return HttpResponse(json.dumps(geojson), content_type='application/json')


@login_required(login_url='/admin/login/')
def forecast(request, code):
    
    selected_barangay = Barangay.objects.get(code=code)

    if not request.user.is_superuser:
        user = UserMunicipal.objects.get(user=request.user)
        
        if user.municipal != selected_barangay.municipal:
            return redirect('total_cases_barangay')
    
     
    cases_by_day = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__barangay=selected_barangay).annotate(date_only=TruncDate('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
    result_day = [{'date': item['date_only'].strftime('%Y-%m-%d'), 'total_cases': item['total_cases']} for item in cases_by_day]
    data_day = json.dumps(result_day) 
    
    cases_by_week = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__barangay=selected_barangay).annotate(date_only=TruncWeek('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
    result_week = [{'date': item['date_only'].strftime('%Y-%m-%d'), 'total_cases': item['total_cases']} for item in cases_by_week]
    data_week = json.dumps(result_week)
    
    cases_by_month = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__barangay=selected_barangay).annotate(month=TruncMonth('date')).values('month').annotate(total_cases=Count('id')).order_by('month')
    result_month = [{'month': item['month'].strftime('%Y-%m'), 'total_cases': item['total_cases']} for item in cases_by_month]
    data_month = json.dumps(result_month)
    
    case_list = Case.objects.filter(resident__barangay=selected_barangay).prefetch_related('resident', 'resident__barangay', 'resident__municipal').all().order_by('-date')
    case_number = case_list.count()

    print("Finish Query...")
    
    df_dengue = pd.DataFrame(result_week, columns=['date', 'total_cases'])
    df_dengue['date'] = pd.to_datetime(df_dengue['date'])
    df_dengue.set_index('date', inplace=True)
    weekly_data = df_dengue.resample('W-MON').sum()
    weekly_data = weekly_data.reindex(pd.date_range(start=f"{df_dengue.index.min().year}-01-01", end=df_dengue.index.max(), freq='W-MON'), fill_value=0)
    weekly_data['total_cases'] = weekly_data['total_cases'].replace({0: 0.3})
    
    
    print(selected_barangay)
    print(weekly_data.count()['total_cases'])
    print(weekly_data)

    print("Finish Trasforming...")

    
    if weekly_data.count()['total_cases'] >= 15:

        
        train_data = weekly_data.copy()
        
        
        if selected_barangay.is_auto_arima:
            model_fit = auto_arima(train_data['total_cases'], stepwise=False, error_action='ignore', n_jobs=-1, seasonal=False, max_p=3, max_d=2, max_q=3, cache=True)
            print("Using auto arima")
            print("Order:", model_fit.order)
            print("Finish Fitting...")
            prediction = model_fit.predict(n_periods=1)  
            warnings.filterwarnings('ignore')
            forecast = {
                "date": prediction.index[0],
                "number_of_cases": round(prediction[0])
            }

        else:
            print("Using manual arima")
            
            if selected_barangay.is_seasonal == True:  
                
                
                model = SARIMAX(train_data['total_cases'], order=(selected_barangay.ar ,selected_barangay.i  ,selected_barangay.ma), seasonal_order=(selected_barangay.seasonal_ar ,selected_barangay.seasonal_i  ,selected_barangay.seasonal_ma, 52))
                warnings.filterwarnings('ignore')
                model_fit = model.fit()
                print('Seasonal')
                prediction = model_fit.forecast(1)  
                warnings.filterwarnings('ignore')
                forecast = {
                    "date": prediction.index[0],
                    "number_of_cases": round(prediction[0])
                }
            
            else:
                model = SARIMAX(train_data['total_cases'], order=(selected_barangay.ar ,selected_barangay.i  ,selected_barangay.ma))
                warnings.filterwarnings('ignore')
                model_fit = model.fit()
                print("Order: ", selected_barangay.ar ,selected_barangay.i  ,selected_barangay.ma)
                print('Not Seasonal')
                prediction = model_fit.forecast(1)  
                warnings.filterwarnings('ignore')
                forecast = {
                    "date": prediction.index[0],
                    "number_of_cases": round(prediction[0])
                }
                    
        
        print(prediction.index[0])
        print(prediction[0])
        
       
       
        
        return render(request, 'dengue/forecast.html',{
        'selected_barangay': selected_barangay,
        'forecast': forecast,
        'prediction_date_plus_6_days': prediction.index[0] + timedelta(days=6),
        'data_day': data_day,
        'data_week': data_week,
        'data_month': data_month,
        'case_list': case_list,
        'case_number': case_number,
    })

    else:

        return render(request, 'dengue/forecast.html',{
            'selected_barangay': selected_barangay,
            'data_day': data_day,
            'data_week': data_week,
            'data_month': data_month,
            'case_list': case_list,
            'case_number': case_number,
        })
   

@login_required(login_url='/admin/login/')
@user_passes_test(superuser_check)
def data_tables(request):
      
    selected_municipal = request.GET.get('municipal')
    selected_barangay = request.GET.get('barangay')
    fromDate = request.GET.get('fromDate')
    toDate = request.GET.get('toDate')
    
    if selected_municipal and selected_barangay and fromDate and toDate:
        start_date = datetime.strptime(fromDate, '%Y-%m-%d')
        end_date = datetime.strptime(toDate, '%Y-%m-%d') + timedelta(days=1)
        if selected_municipal != 'All':
            if selected_barangay == 'All':
                cases = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__municipal__code=selected_municipal).filter(date__gte=start_date, date__lte=end_date).order_by('-date')
                
            else:
                cases = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__barangay__code=selected_barangay).filter(date__gte=start_date, date__lte=end_date).order_by('-date')
        else:
            cases = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(date__gte=start_date, date__lte=end_date).order_by('-date')
        pdf = True

    else:
        currentDate = datetime.now().date() + timedelta(days=1)
        startofYear = datetime(currentDate.year, 1, 1).date()
        
        cases = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(date__gte=startofYear, date__lte=currentDate).order_by('-date')

        fromDate = str(startofYear)
        toDate = str(currentDate - timedelta(days=1))
        
        pdf = False

    
    municipals = Municipal.objects.all()
    
    items_per_page = 12

    paginator = Paginator(cases, items_per_page)

    page_number = request.GET.get('page')

    page = paginator.get_page(page_number)

    return render(request, 'dengue/data_tables.html',{
        'municipals': municipals,
        'fromDate': fromDate,
        'toDate': toDate,
        'page': page,
        'selected_municipal': selected_municipal,
        'selected_barangay': selected_barangay,
        'pdf_available': pdf,
        'total_cases': cases.count()
    })


@login_required(login_url='/admin/login/')
def total_cases_barangay(request):

    if request.method == 'POST':
        timeFrameOption = request.POST.get('timeFrameOption')
        search = request.POST['searched']

        print(timeFrameOption)
        if timeFrameOption == 'last_7_days':  

            date_7_days_ago = datetime.now().date() - timedelta(days=7)  
            cases_per_barangay = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(Q(resident__barangay__barangay__icontains=search) | Q(resident__barangay__municipal__municipal__icontains=search), date__gte=date_7_days_ago).values('resident__barangay__id', 'resident__barangay__barangay','resident__barangay__code','resident__barangay__municipal__municipal').annotate(num_cases=Count('id')).order_by('-num_cases')
            text = 'selected'

        else:
            cases_per_barangay = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(Q(resident__barangay__barangay__icontains=search) | Q(resident__barangay__municipal__municipal__icontains=search)).values('resident__barangay__id', 'resident__barangay__barangay','resident__barangay__code','resident__barangay__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__barangay__municipal__municipal')
            text =''

        return render(request, 'dengue/total_cases_barangay.html',{
            'cases_per_barangay': cases_per_barangay,
            'text': text,
            'search': search,
            'timeFrameOption': timeFrameOption,
        })




    cases_per_barangay = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').values('resident__barangay__id', 'resident__barangay__barangay','resident__barangay__code','resident__barangay__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__barangay__municipal__municipal')



    return render(request, 'dengue/total_cases_barangay.html',{
        'cases_per_barangay': cases_per_barangay,
    })
  
    
@login_required(login_url='/admin/login/')
def total_cases_municipal(request):
   
    if request.method == 'POST':
        selected_option = request.POST.get('time_range')
        
        if selected_option == '7days':
            user = request.user
            if user.is_superuser:

                date_7_days_ago = datetime.now().date() - timedelta(days=7)  
                cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(date__gte=date_7_days_ago).values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__municipal__municipal')
                           
            else: 
                
                try:
                
                    municipal = request.user.usermunicipal.municipal
                    date_7_days_ago = datetime.now().date() - timedelta(days=7)  
                    cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__municipal=municipal, date__gte=date_7_days_ago ).values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__municipal__municipal')

                except ObjectDoesNotExist:
                    cases_per_municipal = ''
            
            return render(request, 'dengue/total_cases_municipal.html',{
                'cases_per_municipal': cases_per_municipal,
                'text': 'Last 7 days'
            
            })
            
        if selected_option == '30days':
            user = request.user
            if user.is_superuser:
                
                date_30_days_ago = datetime.now().date() - timedelta(days=30)  
                cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(date__gte=date_30_days_ago).values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__municipal__municipal')
                
            else: 
                
                try:
                
                    municipal = request.user.usermunicipal.municipal
                    date_30_days_ago = datetime.now().date() - timedelta(days=30)  
                    cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__municipal=municipal, date__gte=date_30_days_ago ).values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__municipal__municipal')
                        
                except ObjectDoesNotExist:
                    cases_per_municipal = ''
            
            return render(request, 'dengue/total_cases_municipal.html',{
                'cases_per_municipal': cases_per_municipal,
                'text': 'Last 30 days'
            
            })
            
        else:
             
            user = request.user
            
            
            if user.is_superuser:

                cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__municipal__municipal')

            else: 
                
                try:
                
                    municipal = request.user.usermunicipal.municipal
                    
                    cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__municipal=municipal).values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__municipal__municipal')
                       
                except ObjectDoesNotExist:
                    cases_per_municipal = ''


            return render(request, 'dengue/total_cases_municipal.html',{
                'cases_per_municipal': cases_per_municipal,
                'text': 'Total'
            
            })
                
    else:
   
        
        user = request.user
        
        
        if user.is_superuser:


            cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__municipal__municipal')
                    
        else: 
            
            try:
            
                municipal = request.user.usermunicipal.municipal
                
                cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__municipal=municipal).values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__municipal__municipal')
                    
            except ObjectDoesNotExist:
                cases_per_municipal = ''

        

    
        return render(request, 'dengue/total_cases_municipal.html',{
            'cases_per_municipal': cases_per_municipal,
            'text': 'Total'
        
        })


def choropleth(request):
    
    if request.method == 'POST':
        
        selected_option = request.POST.get('time_range')
        
        print(selected_option)
        
        if selected_option == "7days":
            
            now = datetime.now().date()
        
            print(now)
            date_7_days_ago = datetime.now().date() - timedelta(days=6)

            print(date_7_days_ago)

            cases_subquery = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(
                resident__barangay=OuterRef('id'),
                date__gte=date_7_days_ago, 
            ).values('resident__barangay__id'
            ).annotate(num_cases=Count('id')
            ).values('num_cases')
            
            cases_per_barangay = Barangay.objects.prefetch_related('resident','resident__barangay','resident__municipal').annotate(
                num_cases=Coalesce(Subquery(cases_subquery, output_field=models.IntegerField()), 0)
            ).values(
                'id',
                'barangay',
                'code',
                'municipal__municipal',
                'geom', 
                'num_cases'
            ).order_by('municipal__municipal')
            
            cases_list = list(cases_per_barangay)
            
            
            filtered_total = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=date_7_days_ago)
            
            total_7 = format(int(filtered_total.count()),',')
    
    
            features = []
            for case in cases_list:
                geom = case['geom']
                if geom:
                    feature = geojson.Feature(
                        geometry=geojson.loads(geom.geojson),
                        properties={
                            'id': case['id'],
                            'barangay': case['barangay'],
                            'code': case['code'],
                            'municipal': case['municipal__municipal'],
                            'num_cases': case['num_cases']
                        }
                    )
                    features.append(feature)

            feature_collection = geojson.FeatureCollection(features)

            json_barangay = geojson.dumps(feature_collection, sort_keys=True)  
            
            return render(request, 'dengue/choropleth.html',{
            'json_barangay': json_barangay,
            'total_7': total_7,
             'text': 'Last 7 days',
        })
            
        elif selected_option == "30days":
            
            now = datetime.now().date()
        
            print(now)
            date_30_days_ago = datetime.now().date() - timedelta(days=30)

            print(date_30_days_ago)

            cases_subquery = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(
                resident__barangay=OuterRef('id'),
                date__gte=date_30_days_ago,
            ).values('resident__barangay__id').annotate(
                num_cases=Count('id')
            ).values('num_cases')

            cases_per_barangay = Barangay.objects.prefetch_related('resident','resident__barangay','resident__municipal').annotate(
                num_cases=Coalesce(Subquery(cases_subquery, output_field=models.IntegerField()), 0)
            ).values(
                'id',
                'barangay',
                'code',
                'municipal__municipal',
                'geom',
                'num_cases'
            ).order_by('municipal__municipal')
            
            
            cases_list = list(cases_per_barangay)
            
        
            filtered_total = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=date_30_days_ago)
            
            total_30 = format(int(filtered_total.count()),',')
    
            
            features = []
            for case in cases_list:
                geom = case['geom']
                if geom:
                    feature = geojson.Feature(
                        geometry=geojson.loads(geom.geojson),
                        properties={
                            'id': case['id'],
                            'barangay': case['barangay'],
                            'code': case['code'],
                            'municipal': case['municipal__municipal'],
                            'num_cases': case['num_cases']
                        }
                    )
                    features.append(feature)

            feature_collection = geojson.FeatureCollection(features)

            json_barangay = geojson.dumps(feature_collection, sort_keys=True)  
            
            return render(request, 'dengue/choropleth.html',{
            'json_barangay': json_barangay,
            'total_30': total_30,
            'text': 'Last 30 days'
        })
            
    else:
            
        cases_subquery = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__barangay=OuterRef('id')).values('resident__barangay__id').annotate(
            num_cases=Count('id')
        ).values('num_cases')

        cases_per_barangay = Barangay.objects.prefetch_related('resident','resident__barangay','resident__municipal').annotate(
            num_cases=Coalesce(Subquery(cases_subquery, output_field=models.IntegerField()), 0)
        ).values(
            'id',
            'barangay',
            'code',
            'municipal__municipal',
            'geom',
            'num_cases'
        ).order_by('municipal__municipal')
        
        cases_list = list(cases_per_barangay)
        
        features = []
        for case in cases_list:
            geom = case['geom']
            if geom:
                feature = geojson.Feature(
                    geometry=geojson.loads(geom.geojson),
                    properties={
                        'id': case['id'],
                        'barangay': case['barangay'],
                        'code': case['code'],
                        'municipal': case['municipal__municipal'],
                        'num_cases': case['num_cases'],
                    }
                )
                features.append(feature)

        feature_collection = geojson.FeatureCollection(features)

        json_barangay = geojson.dumps(feature_collection, sort_keys=True)
        
        total_all = format(int(Case.objects.count()),',')
        
        return render(request, 'dengue/choropleth.html',{
            'json_barangay': json_barangay,
            'total_all': total_all,
            'text': 'Total'
        })
   
   
def home(request):
    
    currentDate = datetime.now().date() + timedelta(days=1)
    startofYear = datetime(currentDate.year, 1, 1).date()
    
    if request.method == 'POST':
        fromDate = request.POST.get('fromDate')
        toDate = request.POST.get('toDate')
        selected_municipal = request.POST.get('municipal')
        
        print(selected_municipal)
        
        if fromDate and toDate:
            
            
            
            start_date = datetime.strptime(fromDate, '%Y-%m-%d')
            end_date = datetime.strptime(toDate, '%Y-%m-%d') + timedelta(days=1)
            
            if selected_municipal != 'All':
                cases = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__municipal__municipal=selected_municipal).filter(date__gte=start_date, date__lt=end_date)

                cases_by_day = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__municipal__municipal=selected_municipal).filter(date__gte=start_date, date__lt=end_date).annotate(date_only=TruncDate('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')

                result_day = [{'date': item['date_only'].strftime('%Y-%m-%d'), 'total_cases': item['total_cases']} for item in cases_by_day]

                data_day= json.dumps(result_day)
                
                
                cases_by_week = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__municipal__municipal=selected_municipal).filter(date__gte=start_date, date__lt=end_date).annotate(date_only=TruncWeek('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
                
                result_week = [{'date': item['date_only'].strftime('%Y-%m-%d'), 'total_cases': item['total_cases']} for item in cases_by_week]

                data_week= json.dumps(result_week)
                

                cases_by_month = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(resident__municipal__municipal=selected_municipal).filter(date__gte=start_date, date__lt=end_date).annotate(month=TruncMonth('date')).values('month').annotate(total_cases=Count('id')).order_by('month')

                result_month = [{'month': item['month'].strftime('%Y-%m'), 'total_cases': item['total_cases']} for item in cases_by_month]

                data_month = json.dumps(result_month)

                cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__municipal__municipal=selected_municipal).filter(date__gte=start_date, date__lt=end_date).values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('-num_cases')

                cases_per_barangay = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(resident__municipal__municipal=selected_municipal).filter(date__gte=start_date, date__lt=end_date).values('resident__barangay__id', 'resident__barangay__barangay','resident__barangay__code','resident__barangay__municipal__municipal').annotate(num_cases=Count('id')).order_by('-num_cases')
            

            else:
                cases = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=start_date, date__lt=end_date)
                
                cases_by_day = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=start_date, date__lt=end_date).annotate(date_only=TruncDate('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')

                result_day = [{'date': item['date_only'].strftime('%Y-%m-%d'), 'total_cases': item['total_cases']} for item in cases_by_day]

                data_day= json.dumps(result_day)
                
                
                cases_by_week = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=start_date, date__lt=end_date).annotate(date_only=TruncWeek('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
                
                result_week = [{'date': item['date_only'].strftime('%Y-%m-%d'), 'total_cases': item['total_cases']} for item in cases_by_week]

                data_week= json.dumps(result_week)
                

                cases_by_month = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=start_date, date__lt=end_date).annotate(month=TruncMonth('date')).values('month').annotate(total_cases=Count('id')).order_by('month')

                result_month = [{'month': item['month'].strftime('%Y-%m'), 'total_cases': item['total_cases']} for item in cases_by_month]

                data_month = json.dumps(result_month)

                cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(date__gte=start_date, date__lt=end_date).values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('-num_cases')

                cases_per_barangay = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(date__gte=start_date, date__lt=end_date).values('resident__barangay__id', 'resident__barangay__barangay','resident__barangay__code','resident__barangay__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__barangay__municipal__municipal')

                

            
            
            features = []

            for case in cases:
                feature = {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [ case.resident.geom.x,case.resident.geom.y],
                    },
                    "properties": {
                        "date": case.date.strftime('%Y-%m-%d'),
                        "dengue_type": case.dengue_type,
                        "resident_id": case.resident.resident_id,
                        "municipal": case.resident.municipal.municipal,
                        "barangay": case.resident.barangay.barangay,
                        "sex": case.resident.sex,
                        "age": currentDate.year - int(case.resident.birth_date.strftime('%Y'))
                    },
                }
                features.append(feature)

            geojson = {
                "type": "FeatureCollection",
                "features": features,
            }
        
        
            all_cases = json.dumps(geojson)

            
            municipals = Municipal.objects.all()


            total_cases = sum(case['num_cases'] for case in cases_per_municipal)
            print(total_cases)

            for case in cases_per_municipal:
                case['percentage'] = (case['num_cases'] * 100) / total_cases
                        
            return render(request, 'dengue/home.html',{
                'municipals': municipals,
                'selected_municipal': selected_municipal,
                'all_cases': all_cases, 
                'fromDate': fromDate,
                'toDate': toDate,   
                'fromDate_datetime': startofYear,
                'toDate_datetime': currentDate - timedelta(days=1),    
                'data_day':data_day,
                'data_week': data_week,
                'data_month': data_month,  
                'cases': cases,
                'cases_per_municipal': cases_per_municipal,
                'cases_per_barangay': cases_per_barangay,
                'total_cases': total_cases,
            })
            
        else:
            return redirect('home')
            
        
   
        
    else:
        
        municipals = Municipal.objects.all()
                
        cases = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=startofYear, date__lte=currentDate)

        features = []

        for case in cases:
            feature = {
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    "coordinates": [ case.resident.geom.x,case.resident.geom.y],
                },
                "properties": {
                    "date": case.date.strftime('%Y-%m-%d'),
                    "dengue_type": case.dengue_type,
                    "resident_id": case.resident.resident_id,
                    "municipal": case.resident.municipal.municipal,
                    "barangay": case.resident.barangay.barangay,
                    "sex": case.resident.sex,
                    "age": currentDate.year - int(case.resident.birth_date.strftime('%Y'))
                },
            }
            features.append(feature)

        geojson = {
            "type": "FeatureCollection",
            "features": features,
        }


        all_cases = json.dumps(geojson)

        
        cases_by_day = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=startofYear, date__lte=currentDate).annotate(date_only=TruncDate('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
        
        result_day = [{'date': item['date_only'].strftime('%Y-%m-%d'), 'total_cases': item['total_cases']} for item in cases_by_day]

        data_day= json.dumps(result_day)
        
        
        cases_by_week = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=startofYear, date__lte=currentDate).annotate(date_only=TruncWeek('date')).values('date_only').annotate(total_cases=Count('id')).order_by('date_only')
        
        result_week = [{'date': item['date_only'].strftime('%Y-%m-%d'), 'total_cases': item['total_cases']} for item in cases_by_week]

        data_week= json.dumps(result_week)
        

        cases_by_month = Case.objects.prefetch_related('resident','resident__barangay','resident__municipal').filter(date__gte=startofYear, date__lte=currentDate).annotate(month=TruncMonth('date')).values('month').annotate(total_cases=Count('id')).order_by('month')

        result_month = [{'month': item['month'].strftime('%Y-%m'), 'total_cases': item['total_cases']} for item in cases_by_month]

        data_month = json.dumps(result_month)


        cases_per_municipal = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(date__gte=startofYear, date__lte=currentDate).values('resident__municipal__id', 'resident__municipal__municipal').annotate(num_cases=Count('id')).order_by('-num_cases')

        cases_per_barangay = Case.objects.prefetch_related('resident', 'resident__barangay', 'resident__municipal').filter(date__gte=startofYear, date__lte=currentDate).values('resident__barangay__id', 'resident__barangay__barangay','resident__barangay__code','resident__barangay__municipal__municipal').annotate(num_cases=Count('id')).order_by('resident__barangay__municipal__municipal')

        total_cases = sum(case['num_cases'] for case in cases_per_municipal)
        print(total_cases)

        for case in cases_per_municipal:
            case['percentage'] = (case['num_cases'] * 100) / total_cases

        return render(request, 'dengue/home.html',{
            'municipals': municipals,
            'data_day':data_day,
            'data_week': data_week,
            'data_month': data_month,
            'all_cases': all_cases, 
            'fromDate': str(startofYear),
            'toDate': str(currentDate - timedelta(days=1)),    
            'fromDate_datetime': startofYear,
            'toDate_datetime': currentDate - timedelta(days=1),   
            'cases': cases,
            'cases_per_municipal': cases_per_municipal,
            'cases_per_barangay': cases_per_barangay,
            'total_cases': total_cases,
        })
